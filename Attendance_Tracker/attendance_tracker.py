import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# To read and update the Excel sheet
def load_workbook(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    return workbook, sheet

# Function to send email
def send_email(to_email, subject, message_body):
    sender_email = "daudivincent20@gmail.com"
    sender_password = "butz gsdi hbpj uaop"  # Use an app-specific password if using Gmail

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = to_email
    msg['Subject'] = subject

    body = MIMEText(message_body, 'plain')
    msg.attach(body)

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, to_email, msg.as_string())
        server.quit()
        print(f"Email sent to {to_email}")
    except Exception as e:
        print(f"Failed to send email: {e}")

# Function to read Excel file and track attendance
def track_attendance(excel_file, roll_number, subject_code):
    workbook, sheet = load_workbook(excel_file)

    # Map subject code to corresponding columns
    subject_column = None
    staff_email_column = None
    if subject_code == 'Subject1':
        subject_column = 3  # Subject1_Leaves
        staff_email_column = 6  # StaffEmail_Subject1
    elif subject_code == 'Subject2':
        subject_column = 4
        staff_email_column = 7
    elif subject_code == 'Subject3':
        subject_column = 5
        staff_email_column = 8

    # Iterate through rows to find the student by roll number
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
        if row[0].value == roll_number:
            email = row[1].value
            leave_count = row[subject_column - 1].value
            staff_email = row[staff_email_column - 1].value

            # If student has taken 2 leaves, send a reminder
            if leave_count == 2:
                send_email(email, f"Reminder: Attendance for {subject_code}",
                           f"Dear student, you have taken 2 leaves for {subject_code}. Only 1 more leave is allowed.")
            # If student has taken 3 or more leaves, notify both student and staff
            elif leave_count >= 3:
                send_email(email, f"Attendance Warning for {subject_code}",
                           f"Dear student, you have exceeded the allowed leaves for {subject_code}.")
                send_email(staff_email, f"Attendance Alert for {subject_code}",
                           f"Student {roll_number} has exceeded the allowed leaves for {subject_code}.")

            # Increment leave count and save the updated workbook
            row[subject_column - 1].value = leave_count + 1
            workbook.save(excel_file)
            print(f"Attendance updated for {roll_number} in {subject_code}")
            break

# Example usage: 
track_attendance('attendance.xlsx', '101', 'Subject1')
